import streamlit as st
import datetime
import pandas as pd
from supabase import create_client, Client
import json
import os

# Load Supabase credentials (works locally and on Streamlit Cloud)
try:
    SUPABASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
except Exception:
    from dotenv import load_dotenv
    load_dotenv()
    SUPABASE_URL = os.getenv("SUPABASE_URL")
    SUPABASE_KEY = os.getenv("SUPABASE_KEY")


@st.cache_data
def load_excel_data():
    """Loads and parses the assets.xlsx file to build mapping dictionaries."""
    zone_mapping = {}
    concern_persons_by_zone = {}
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook('assets.xlsx', data_only=True)
        
        # Parse Site List sheet for Zone -> Site Name -> UnitCode
        if 'Site List' in wb.sheetnames:
            ws_sites = wb['Site List']
            for row in ws_sites.iter_rows(min_row=2, values_only=True): # Skip header
                if not row or row[0] is None: continue
                # Explicitly cast to string and strip so integer 0 becomes "0"
                zone = str(row[0]).strip()
                unit_code = row[1]
                site_name = row[2]
                
                if zone not in zone_mapping:
                    zone_mapping[zone] = {}
                
                # If site_name is empty, ignore it 
                if site_name:
                    zone_mapping[zone][site_name] = unit_code
                    
        # Parse Branch Team sheet for concern persons mapped by zone
        if 'Branch Team' in wb.sheetnames:
            ws_team = wb['Branch Team']
            for row in ws_team.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None or row[1] is None: continue 
                zone = str(row[0]).strip()
                name = str(row[1]).strip()
                
                if zone not in concern_persons_by_zone:
                    concern_persons_by_zone[zone] = []
                    
                if name and name not in concern_persons_by_zone[zone]:
                    concern_persons_by_zone[zone].append(name)
            
            # Ensure "Not Required" is explicitly the first option for every zone
            for z in concern_persons_by_zone:
                if "Not Required" in concern_persons_by_zone[z]:
                    concern_persons_by_zone[z].remove("Not Required")
                concern_persons_by_zone[z].insert(0, "Not Required")
                    
    except Exception as e:
        st.error(f"Failed to load assets.xlsx: {e}")
        # Fallback dummy data if file fails
        zone_mapping = {1: {"Error Loading Data": 0}}
        concern_persons_by_zone = {1: ["Not Required", "Error Loading Personnel"]}
        
    return zone_mapping, concern_persons_by_zone

# Load the dynamic data
ZONE_MAPPING, CONCERN_PERSONS_BY_ZONE = load_excel_data()

# Set page config
try:
    st.set_page_config(page_title="SGV Ops Desk", page_icon="favicon.ico", layout="wide", initial_sidebar_state="expanded")
except:
    st.set_page_config(page_title="SGV Ops Desk", page_icon="🏢", layout="wide", initial_sidebar_state="expanded")

# Minimalist & Modern CSS Styling
st.markdown("""
<style>
    /* Hide Streamlit Default Elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Hide Deploy button but KEEP sidebar toggle */
    .stDeployButton {display: none;}

    /* Main Background & Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
    
    .stApp {
        background-color: #F8FAFC;
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    
    /* Global Text Styles */
    * {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
    }
    h1, h2, h3 {
        color: #0F172A !important;
        font-weight: 700 !important;
        letter-spacing: -0.02em;
    }
    
    /* Input Fields (Minimalist) */
    .stTextInput>div>div>input, 
    .stNumberInput>div>div>input, 
    .stSelectbox>div>div>div, 
    .stTextArea>div>div>textarea,
    .stTimeInput>div>div>input,
    .stDateInput>div>div>input,
    .stMultiSelect>div>div>div {
        border-radius: 6px !important;
        border: 1px solid #E2E8F0 !important;
        background-color: #FFFFFF !important;
        padding-left: 14px;
        font-size: 14px !important;
        color: #1E293B !important;
        box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease-in-out;
    }
    
    .stTextInput>div>div>input:focus, 
    .stNumberInput>div>div>input:focus,
    .stSelectbox>div>div>div:focus, 
    .stTextArea>div>div>textarea:focus,
    .stMultiSelect>div>div>div:focus {
        border-color: #0F172A !important;
        box-shadow: 0 0 0 1px #0F172A !important;
    }
    
    /* Hide number input arrows */
    [data-testid="stNumberInputStepUp"],
    [data-testid="stNumberInputStepDown"] {
        display: none !important;
    }
    input[type=number]::-webkit-inner-spin-button, 
    input[type=number]::-webkit-outer-spin-button { 
        -webkit-appearance: none; 
        margin: 0; 
    }
    input[type=number] {
        -moz-appearance: textfield;
    }
    
    /* Label text */
    .stTextInput label p, .stSelectbox label p, .stDateInput label p, .stTimeInput label p, .stNumberInput label p, .stMultiSelect label p, .stTextArea label p {
        font-size: 13px !important;
        font-weight: 600 !important;
        color: #475569 !important;
        margin-bottom: 4px !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    /* Form Container (Clean Card) */
    [data-testid="stForm"] {
        background-color: #FFFFFF;
        padding: 40px;
        border-radius: 16px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -2px rgba(0, 0, 0, 0.05);
    }
    
    /* Primary Buttons (Sleek Dark) */
    .stButton>button, 
    [data-testid="stFormSubmitButton"]>button {
        background-color: #0F172A !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        border-radius: 6px !important;
        padding: 12px 28px !important;
        border: none !important;
        font-size: 14px !important;
        box-shadow: 0 4px 6px -1px rgba(15, 23, 42, 0.2) !important;
        transition: all 0.2s ease !important;
        width: auto !important;
    }
    
    .stButton>button:hover,
    [data-testid="stFormSubmitButton"]>button:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 12px -2px rgba(15, 23, 42, 0.3) !important;
        background-color: #1E293B !important;
    }
    
    /* Tabs (Minimalist under-line) */
    .stTabs [data-baseweb="tab-list"] {
        gap: 32px;
        border-bottom: 1px solid #E2E8F0;
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        height: 48px;
        white-space: pre-wrap;
        background-color: transparent;
        border-radius: 0;
        gap: 1px;
        padding: 0 4px;
        color: #64748b;
        font-weight: 600;
        font-size: 15px;
    }
    .stTabs [aria-selected="true"] {
        color: #0F172A;
        border-bottom: 2px solid #0F172A;
        background-color: transparent !important;
    }
    
    /* MultiSelect Tags */
    .stMultiSelect [data-baseweb="tag"] {
        background-color: #F1F5F9;
        border-radius: 4px;
        color: #0F172A;
    }
</style>
""", unsafe_allow_html=True)


@st.cache_resource
def get_supabase_client():
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    try:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        return supabase
    except Exception as e:
        st.error(f"Failed to connect to Supabase: {e}")
        return None

def check_password():
    """Returns `True` if the user had a correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        try:
            with open("users.json", "r") as f:
                users_records = json.load(f)
        except Exception as e:
            st.error(f"Could not load users.json: {e}")
            st.session_state["password_correct"] = False
            return
            
        valid_user = None
        for u in users_records:
            user_id = str(u.get("UserId", "")).strip()
            name = str(u.get("Name", "")).strip()
            if user_id == st.session_state["username"] or name.lower() == st.session_state["username"].lower():
                valid_user = u
                break

        if valid_user and str(valid_user.get("Password", "")) == st.session_state["password"]:
            st.session_state["password_correct"] = True
            st.session_state["role"] = valid_user.get("Role", "")
            st.session_state["user_id"] = valid_user.get("UserId", "")
            st.session_state["name"] = valid_user.get("Name", "")
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show inputs for username + password.
        st.title("Login to Ops Desk")
        st.info("Please login with your UserId or Name.")
        st.text_input("Username / UserId", key="username")
        st.text_input("Password", type="password", key="password")
        st.button("Login", on_click=password_entered)
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.title("Login to Ops Desk")
        st.text_input("Username / UserId", key="username")
        st.text_input("Password", type="password", key="password")
        st.button("Login", on_click=password_entered)
        st.error("😕 User not known or password incorrect")
        return False
    else:
        # Password correct.
        return True

def generate_case_number(supabase: Client):
    """Generates next case number: SGV_CR_YYYY_MM_X and keeps incrementing X"""
    now = datetime.datetime.now()
    year_month = now.strftime("%Y_%m")
    prefix_current = f"SGV_CR_{year_month}_"
    
    try:
        # Fetch the most recently created case that starts with this month's prefix
        # We order by Timestamp descending and limit to 1
        response = supabase.table("cases")\
            .select("CaseNumber")\
            .like("CaseNumber", f"{prefix_current}%")\
            .order("Timestamp", desc=True)\
            .limit(1)\
            .execute()
            
        if not response.data or len(response.data) == 0:
            return f"{prefix_current}1"
            
        last_case = response.data[0]["CaseNumber"]
        
        if last_case and last_case.startswith(prefix_current):
            # Extract the number part
            parts = last_case.split("_")
            if len(parts) >= 5:
                last_count = parts[-1]
                if last_count.isdigit():
                    next_count = int(last_count) + 1
                    return f"{prefix_current}{next_count}"
                    
        return f"{prefix_current}1"
    except Exception as e:
        # If the table doesn't exist or other error, fallback to 1
        st.warning(f"Warning fetching last case number: {e}")
        return f"{prefix_current}1"


# --- MAIN APP ---
supabase = get_supabase_client()

if not supabase:
    st.title("Ops Desk - Database Setup Required")
    st.error("Missing Supabase credentials.")
    st.info("Please set `SUPABASE_URL` and `SUPABASE_KEY` in the `.env` file in this directory.")
else:
    if check_password():
        # --- MAIN HEADER ---
        col_logo, col_title, col_user = st.columns([1, 6, 2])
        
        with col_logo:
            try:
                st.image("logo.png", width='stretch')
            except Exception:
                st.markdown("<h2>SGV</h2>", unsafe_allow_html=True)
                
        with col_title:
            st.title("Operations Desk")
            st.markdown("<p style='color: #64748b; font-size: 15px; margin-top: -15px;'>Manage and log all your operational calls efficiently.</p>", unsafe_allow_html=True)
            
        with col_user:
            st.markdown(f"""
            <div style="background-color: #F1F5F9; padding: 12px; border-radius: 8px; border: 1px solid #E2E8F0; margin-bottom: 10px;">
                <p style="margin:0; font-size:11px; color:#64748b; font-weight:600; text-transform:uppercase;">Logged In As</p>
                <p style="margin:0; font-size:14px; color:#0F172A; font-weight:700;">{st.session_state['name']}</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Log Out", width='stretch'):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()

        st.markdown("<hr style='margin-top: 5px; margin-bottom: 25px;'>", unsafe_allow_html=True)
        
        # --- TABS ---
        tab1, tab2, tab3 = st.tabs(["📋 Log New Case", "🔄 Update Case", "📊 View & Export Data"])
        
        with tab1:
            st.markdown("<br>", unsafe_allow_html=True)
            call_type_selection = st.selectbox(
                "SELECT CALL TYPE",
                options=["", "Incoming", "Outgoing"],
                index=0
            )
            
            if call_type_selection:
                st.subheader(f"{call_type_selection} Call Details")
                
                # Cascading dropdowns (must be outside the form to trigger live updates)
                st.markdown("### Location Details")
                col_x, col_y, col_z = st.columns(3)
                
                with col_x:
                    zone = st.selectbox("ZONE", options=list(ZONE_MAPPING.keys()))
                
                sites_for_zone = ZONE_MAPPING.get(zone, {})
                
                with col_y:
                    site_name = st.selectbox("SITE NAME", options=list(sites_for_zone.keys()))
                
                unitcode = sites_for_zone.get(site_name, 0)
                
                with col_z:
                    st.number_input("UNIT CODE", value=unitcode, disabled=True)
                
                st.markdown("---")
                
                with st.form(key="call_details_form", clear_on_submit=True):
                    st.info("Case Number will be auto-generated on submission.")
                    
                    # Row 1
                    col1, col2 = st.columns(2)
                    with col1:
                        date = st.date_input("DATE", datetime.date.today())
                    with col2:
                        time = st.time_input("TIME", datetime.datetime.now().time())
                        
                    # Row 2
                    col3, col4 = st.columns(2)
                    with col3:
                        emp_id_input = st.text_input("EMP.ID (Numeric)")
                    with col4:
                        emp_name = st.text_input("EMP.NAME")
                        
                    # Row 3
                    col5, col6 = st.columns(2)
                    with col5:
                        case_type_selection = st.selectbox("CASE TYPE", ["Reporting", "Incident", "Fire", "Medical", "Security", "Other"])
                        if case_type_selection == "Other":
                            case_type_custom = st.text_input("SPECIFY OTHER CASE TYPE")
                        else:
                            case_type_custom = ""
                    with col6:
                        # Use time_input but allow None (Streamlit 1.28+) or just default to now. 
                        # We use time_input to strictly prevent user formatting bugs like "2pm" or "abc"
                        reported_time = st.time_input("REPORTED TIME", value=None)
                        
                    # Row 4
                    col7, col8 = st.columns(2)
                    with col7:
                        status = st.selectbox("STATUS", ["Open", "Closed", "No Resolution"])
                    with col8:
                        concern_persons_for_zone = CONCERN_PERSONS_BY_ZONE.get(zone, [])
                        concern_person = st.multiselect("CONCERN PERSON", options=concern_persons_for_zone)
                        
                    # Row 5 (Full Width to prevent breaking column heights)
                    purpose_problem = st.text_area("PURPOSE / PROBLEM", height=100)
                    remarks = st.text_area("REMARKS", height=100)
                    
                    submit_button = st.form_submit_button(label="Submit Record")
                    
                    if submit_button:
                        # Validation for Employee ID and Custom Case Type
                        is_valid = True
                        emp_id_val = None
                        
                        if emp_id_input:
                            if not emp_id_input.isdigit():
                                st.error("EMP.ID must be a pure number.")
                                is_valid = False
                            else:
                                emp_id_val = int(emp_id_input)
                                
                        if case_type_selection == "Other" and not case_type_custom.strip():
                            st.error("Please specify the custom Case Type.")
                            is_valid = False
                            
                        if is_valid:
                            with st.spinner("Generating Case Number and saving to Supabase..."):
                                new_case_number = generate_case_number(supabase)
                                
                                # Resolve case type
                                final_case_type = case_type_custom.strip() if case_type_selection == "Other" else case_type_selection
                                
                                # Format date and time safely (DD-MM-YYYY for easy reading)
                                date_str = date.strftime("%d-%m-%Y") if date else ""
                                time_str = time.strftime("%H:%M:%S") if time else ""
                                
                                # Override reported time if Not Required
                                if "Not Required" in concern_person:
                                    reported_time_str = "00:00:00"
                                else:
                                    reported_time_str = reported_time.strftime("%H:%M:%S") if reported_time else ""
                                
                                # Supabase timestamp format (ISO 8601)
                                timestamp_iso = datetime.datetime.now().isoformat()
                                
                                # Prepare dictionary data for Supabase insert
                                row_data = {
                                    "CaseNumber": new_case_number,
                                    "DATE": date_str,
                                    "TIME": time_str,
                                    "EMP_ID": emp_id_val,
                                    "EMP_NAME": emp_name,
                                    "ZONE": zone,
                                    "UNITCODE": unitcode,
                                    "SITE_NAME": site_name,
                                    "CASE_TYPE": final_case_type,
                                    "PURPOSE_PROBLEM": purpose_problem,
                                    "CALL_TYPE": call_type_selection,
                                    "STATUS": status,
                                    "REMARKS": remarks,
                                    "CONCERN_PERSON": concern_person,  # This is a list/array
                                    "Reported_TIME": reported_time_str,
                                    "Logged_By": st.session_state["name"],
                                    "Timestamp": timestamp_iso
                                }
                            
                            try:
                                response = supabase.table("cases").insert(row_data).execute()
                                st.success(f"Successfully recorded {call_type_selection} call. Case Number: **{new_case_number}**")
                            except Exception as e:
                                st.error(f"Failed to save record to Supabase: {e}")
        
        with tab2:
            st.header("Update Case Status")
            
            search_case = st.text_input("Enter Case Number to Search (e.g., SGV_CR_2026_03_1)")
            
            if st.button("Search Case"):
                if search_case:
                    with st.spinner("Searching..."):
                        try:
                            response = supabase.table("cases").select("*").eq("CaseNumber", search_case).execute()
                            
                            if response.data and len(response.data) > 0:
                                record_data = response.data[0]
                                st.session_state["record_data"] = record_data
                                st.success("Case Found!")
                            else:
                                st.error("Case not found.")
                                if "record_data" in st.session_state:
                                    del st.session_state["record_data"]
                        except Exception as e:
                            st.error(f"Error searching Supabase: {e}")
                else:
                    st.warning("Please enter a case number first.")
            
            # If a case is currently loaded into session
            if "record_data" in st.session_state:
                record_data = st.session_state["record_data"]
                st.markdown("---")
                st.subheader(f"Updating: {record_data.get('CaseNumber', 'Unknown')}")
                
                # Read-only info
                st.write(f"**Date:** {record_data.get('DATE', '')} | **Time:** {record_data.get('TIME', '')}")
                st.write(f"**Employee:** {record_data.get('EMP_NAME', '')} ({record_data.get('EMP_ID', '')}) | **Site:** {record_data.get('SITE_NAME', '')}")
                st.write(f"**Call Type:** {record_data.get('CALL_TYPE', '')}")
                st.text_area("Purpose / Problem (Read-only)", value=record_data.get('PURPOSE_PROBLEM', ''), height=100, disabled=True)
                
                # Updateable fields
                with st.form("update_case_form"):
                        
                    current_status = record_data.get('STATUS', 'Open')
                    status_options = ["Open", "Closed", "No Resolution"]
                    default_status_idx = status_options.index(current_status) if current_status in status_options else 0
                    
                    new_status = st.selectbox("STATUS", status_options, index=default_status_idx)
                    new_remarks = st.text_area("REMARKS", value=record_data.get('REMARKS', ''), height=100)
                    
                    update_btn = st.form_submit_button("Update Case")
                    
                    if update_btn:
                        case_id = record_data.get('CaseNumber')
                        if case_id:
                            with st.spinner("Updating Supabase..."):
                                try:
                                    # Update specific row in Supabase
                                    update_data = {
                                        "STATUS": new_status,
                                        "REMARKS": new_remarks
                                    }
                                    response = supabase.table("cases").update(update_data).eq("CaseNumber", case_id).execute()
                                    
                                    # Update session state to reflect new values locally
                                    st.session_state["record_data"]["STATUS"] = new_status
                                    st.session_state["record_data"]["REMARKS"] = new_remarks
                                    
                                    st.success(f"Case {case_id} updated successfully!")
                                except Exception as e:
                                    st.error(f"Failed to update Supabase: {e}")

        with tab3:
            st.header("View and Export Call Records")
            st.markdown("Filter records by date and zone, then download them as Excel or PDF.")
            
            col_d1, col_d2, col_d3, col_d4 = st.columns(4)
            with col_d1:
                # Default start date is the 1st of the current month
                start_date = st.date_input("Start Date", datetime.date.today().replace(day=1))
            with col_d2:
                end_date = st.date_input("End Date", datetime.date.today())
            with col_d3:
                zone_filter = st.selectbox("Zone Filter", ["All"] + list(ZONE_MAPPING.keys()))
            with col_d4:
                call_type_filter = st.selectbox("Call Type Filter", ["All", "Inbound", "Outbound"])
                
            if st.button("Fetch Records"):
                with st.spinner("Fetching data from Supabase..."):
                    try:
                        # Convert selected dates to string format matching the database
                        start_str = start_date.strftime("%d-%m-%Y")
                        end_str = end_date.strftime("%d-%m-%Y")
                        
                        # Fetch all records, we will filter robustly in Pandas to handle mixed CSV formats
                        response = supabase.table("cases")\
                            .select("*")\
                            .order("Timestamp", desc=True)\
                            .execute()
                            
                        if not response.data or len(response.data) == 0:
                            st.warning("No records found in the database.")
                        else:
                            # Convert to Pandas DataFrame
                            df = pd.DataFrame(response.data)
                            if 'EMP_ID' in df.columns:
                                df['EMP_ID'] = df['EMP_ID'].apply(lambda x: int(x) if pd.notnull(x) else "")
                            
                            # --- ROBUST DATE FILTERING ---
                            if 'DATE' in df.columns:
                                # Parse dates safely (handles both YYYY-MM-DD and DD-MM-YYYY from manual CSV uploads)
                                df['parsed_date'] = pd.to_datetime(df['DATE'], format='mixed', dayfirst=True, errors='coerce')
                                
                                # Filter the dataframe
                                mask = (df['parsed_date'].dt.date >= start_date) & (df['parsed_date'].dt.date <= end_date)
                                df = df[mask]
                                
                                # Remove temporary column
                                df = df.drop(columns=['parsed_date'])
                                
                            if len(df) == 0:
                                st.warning(f"No records found between {start_str} and {end_str}.")
                                st.stop()
                                
                            # Sort by Timestamp descending to ensure latest cases are naturally at the top
                            if 'Timestamp' in df.columns:
                                # First convert to proper datetime objects, then sort
                                df['Timestamp'] = pd.to_datetime(df['Timestamp'], format='mixed', errors='coerce')
                                df = df.sort_values(by='Timestamp', ascending=False)
                                
                            # Clean up lists in CONCERN_PERSON array for Excel
                            if 'CONCERN_PERSON' in df.columns:
                                df['CONCERN_PERSON'] = df['CONCERN_PERSON'].apply(
                                    lambda x: ", ".join(x) if isinstance(x, list) else str(x)
                                )
                            
                            # Drop the Timestamp column for a cleaner export
                            if 'Timestamp' in df.columns:
                                df = df.drop(columns=['Timestamp'])
                                
                            # Reorder columns to put CASE_TYPE near CALL_TYPE
                            cols = df.columns.tolist()
                            if 'CASE_TYPE' in cols and 'CALL_TYPE' in cols:
                                cols.insert(cols.index('CALL_TYPE') + 1, cols.pop(cols.index('CASE_TYPE')))
                                df = df[cols]
                                
                            # Apply Zone Filter in Pandas
                            if zone_filter != "All":
                                df = df[df['ZONE'] == zone_filter]
                                
                            # Apply Call Type Filter in Pandas
                            if call_type_filter != "All":
                                if 'CALL_TYPE' in df.columns:
                                    # Use robust string matching to ignore case and trailing spaces from CSVs
                                    df = df[df['CALL_TYPE'].astype(str).str.strip().str.lower() == call_type_filter.lower()]
                                
                            if len(df) == 0:
                                st.warning(f"No records found for the selected filters in this date range.")
                            else:
                                st.subheader(f"Found {len(df)} Records:")
                                
                                # --- 1. DISPLAY KPIs ---
                                status_counts = df['STATUS'].value_counts()
                                case_types = df['CASE_TYPE'].value_counts() if 'CASE_TYPE' in df.columns else {}
                                
                                col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)
                                col_kpi1.metric("Total Cases", len(df))
                                col_kpi2.metric("Open", status_counts.get("Open", 0))
                                col_kpi3.metric("Closed", status_counts.get("Closed", 0))
                                col_kpi4.metric("No Resolution", status_counts.get("No Resolution", 0))
                                
                                st.dataframe(df, use_container_width=True)
                                
                                col_btn1, col_btn2 = st.columns(2)
                                
                                # --- 2. EXCEL EXPORT ---
                                import io
                                # Generate Excel file in memory with xlsxwriter to beautify
                                towrite = io.BytesIO()
                                with pd.ExcelWriter(towrite, engine='xlsxwriter') as writer:
                                    df.to_excel(writer, index=False, sheet_name='Call Logs')
                                    
                                    workbook = writer.book
                                    worksheet = writer.sheets['Call Logs']
                                    
                                    # Define beautiful formats
                                    header_format = workbook.add_format({
                                        'bold': True,
                                        'text_wrap': True,
                                        'valign': 'top',
                                        'fg_color': '#1E293B',  # Dark Blue/Slate
                                        'font_color': '#FFFFFF', # White text
                                        'border': 1,
                                        'border_color': '#E2E8F0'
                                    })
                                    
                                    cell_format = workbook.add_format({
                                        'valign': 'top',
                                        'text_wrap': True,
                                        'border': 1,
                                        'border_color': '#E2E8F0'
                                    })
                                    
                                    # Write the column headers with the defined format
                                    for col_num, value in enumerate(df.columns.values):
                                        worksheet.write(0, col_num, value, header_format)
                                        
                                    # Auto-fit columns and apply cell format
                                    for idx, col in enumerate(df):
                                        # Write all row cells with the cell_format
                                        for row_num, val in enumerate(df[col], start=1):
                                            # Convert lists/None to strings safely for xlsxwriter
                                            if val is None:
                                                write_val = ""
                                            elif isinstance(val, list):
                                                write_val = ", ".join(str(v) for v in val)
                                            else:
                                                write_val = str(val)
                                            worksheet.write(row_num, idx, write_val, cell_format)
                                        
                                        # Calculate width based on max data length vs column name
                                        # cap at 50 to prevent huge text areas from making columns unreadable
                                        series = df[col]
                                        max_len = max((
                                            series.astype(str).map(len).max(),
                                            len(str(series.name))
                                        )) + 3
                                        max_len = min(max_len, 50) 
                                        worksheet.set_column(idx, idx, max_len)
                                        
                                towrite.seek(0)
                                
                                with col_btn1:
                                    st.download_button(
                                        label="⬇️ Download Beautified Excel",
                                        data=towrite,
                                        file_name=f"SGV_Ops_Export_{start_str}_to_{end_str}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )
                                    
                                # --- 3. PDF EXPORT ---
                                from reportlab.lib import colors
                                from reportlab.lib.pagesizes import landscape, A4
                                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                                import os
                                
                                pdf_buffer = io.BytesIO()
                                doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                                elements = []
                                styles = getSampleStyleSheet()
                                
                                # Title & Logo Table
                                title_style = ParagraphStyle(name="ReportTitle", parent=styles['Heading1'], alignment=0, spaceAfter=0)
                                title_p = Paragraph("SGV Control Room Report", title_style)
                                
                                logo_item = ""
                                if os.path.exists("logo.png"):
                                    from reportlab.lib.utils import ImageReader
                                    img_reader = ImageReader("logo.png")
                                    img_w, img_h = img_reader.getSize()
                                    target_height = 45 # Perfect height for header
                                    target_width = target_height * (img_w / float(img_h))
                                    logo_item = Image("logo.png", width=target_width, height=target_height)
                                
                                header_table = Table([[title_p, logo_item]], colWidths=[550, 200])
                                header_table.setStyle(TableStyle([
                                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'), 
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                                ]))
                                elements.append(header_table)
                                elements.append(Spacer(1, 15))
                                
                                # Date Processing
                                if start_str == end_str:
                                    date_text = f"<b>Date:</b> {start_str}"
                                else:
                                    date_text = f"<b>Date Range:</b> {start_str} to {end_str}"
                                    
                                # Zone & Call Type Processing
                                zone_text = f"<b>Zone:</b> {zone_filter}" if zone_filter != "All" else "<b>All Zones</b>"
                                call_type_text = f"<b>| Call Type:</b> {call_type_filter}" if call_type_filter != "All" else "<b>| Call Type:</b> All"
                                combined_filter_text = f"{zone_text} &nbsp;&nbsp; {call_type_text}"
                                right_align_style = ParagraphStyle(name="RightAlign", parent=styles['Normal'], alignment=2)
                                
                                # Date (Left) and Filters (Right) Layout
                                date_zone_table = Table([[Paragraph(date_text, styles['Normal']), Paragraph(combined_filter_text, right_align_style)]], colWidths=[375, 375])
                                date_zone_table.setStyle(TableStyle([
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                                ]))
                                elements.append(date_zone_table)
                                elements.append(Spacer(1, 20))
                                
                                # Summary Header
                                elements.append(Paragraph(f"<b>Summary:</b>", styles['Heading3']))
                                elements.append(Spacer(1, 5))
                                
                                kpi_text = f"Total Cases: {len(df)} &nbsp;&nbsp;|&nbsp;&nbsp; Open: {status_counts.get('Open', 0)} &nbsp;&nbsp;|&nbsp;&nbsp; Closed: {status_counts.get('Closed', 0)} &nbsp;&nbsp;|&nbsp;&nbsp; No Resolution: {status_counts.get('No Resolution', 0)}"
                                elements.append(Paragraph(kpi_text, styles['Normal']))
                                
                                elements.append(Spacer(1, 5)) # Gap for readability
                                
                                # Case Type Summary
                                if not case_types.empty:
                                    type_strings = [f"{k}: {v}" for k, v in case_types.items()]
                                    elements.append(Paragraph(f"By Type: {', '.join(type_strings)}", styles['Normal']))
                                
                                elements.append(Spacer(1, 25))
                                
                                # Table Data Prep (Selecting specific columns to fit on PDF)
                                pdf_df = pd.DataFrame()
                                pdf_source_df = df.reset_index(drop=True)
                                
                                pdf_df['S.No'] = range(1, len(pdf_source_df) + 1)
                                pdf_df['Case#'] = pdf_source_df['CaseNumber'].astype(str).str.replace('SGV_CR_', '')
                                pdf_df['Date'] = pdf_source_df['DATE']
                                pdf_df['Site'] = pdf_source_df['SITE_NAME']
                                pdf_df['Type'] = pdf_source_df['CASE_TYPE'] if 'CASE_TYPE' in pdf_source_df.columns else ""
                                pdf_df['Details'] = pdf_source_df['PURPOSE_PROBLEM']
                                pdf_df['Remarks'] = pdf_source_df['REMARKS']
                                pdf_df['Status'] = pdf_source_df['STATUS']
                                
                                data_raw = [pdf_df.columns.to_list()] + pdf_df.astype(str).values.tolist()
                                
                                # Wrap text in Paragraphs to prevent overflow on wide text
                                cell_style = ParagraphStyle(name="Cell", parent=styles["Normal"], fontSize=8, leading=10)
                                wrapped_data = []
                                
                                for row_idx, row in enumerate(data_raw):
                                    if row_idx == 0:
                                        wrapped_data.append(row) # Keep Headers as pure strings
                                    else:
                                        wrapped_row = []
                                        for col_idx, cell_val in enumerate(row):
                                            # Apply text wrapping to Details (5) and Remarks (6)
                                            if col_idx in [5, 6] and cell_val and cell_val != "None" and cell_val != "nan":
                                                wrapped_row.append(Paragraph(cell_val, cell_style))
                                            else:
                                                wrapped_row.append(cell_val if cell_val != "None" and cell_val != "nan" else "")
                                        wrapped_data.append(wrapped_row)
                                
                                # Dynamic Column Widths for PDF (Split available width approx 780 pts)
                                col_widths = [25, 65, 55, 90, 60, 220, 200, 60]
                                
                                t = Table(wrapped_data, colWidths=col_widths, repeatRows=1)
                                t.setStyle(TableStyle([
                                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                                    ('FONTSIZE', (0,0), (-1,-1), 8),
                                    ('BOTTOMPADDING', (0,0), (-1,0), 8),
                                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
                                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
                                    ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
                                ]))
                                elements.append(t)
                                
                                doc.build(elements)
                                pdf_buffer.seek(0)
                                
                                with col_btn2:
                                    st.download_button(
                                        label="📄 Download PDF Management Report",
                                        data=pdf_buffer,
                                        file_name=f"SGV_Mgmt_Report_{start_str}_to_{end_str}.pdf",
                                        mime="application/pdf",
                                        use_container_width=True
                                    )
                            
                    except Exception as e:
                        st.error(f"Failed to fetch or process records: {e}")
